import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import os

def create_comprehensive_grocery_sheet():
    """
    Create a comprehensive Excel sheet with top USA grocery products and brands
    organized by category, using the template structure.
    """

    # Define categories and their top brands/products (50+ for each category)
    categories_data = {
        "Beverages": [
            "Coca-Cola", "Pepsi", "Diet Coke", "Sprite", "Mountain Dew",
            "Dr Pepper", "Gatorade", "Tropicana Orange Juice", "Simply Orange",
            "Minute Maid", "Dasani Water", "Aquafina", "Nestle Pure Life",
            "Smartwater", "Vitaminwater", "Powerade", "Red Bull", "Monster Energy",
            "Starbucks Coffee", "Folgers Coffee", "Maxwell House", "Nescafe",
            "Keurig", "Peet's Coffee", "Lipton Tea", "Arizona Iced Tea",
            "Snapple", "Twinings", "Bigelow Tea", "Celestial Seasonings",
            "Yerba Mate", "Bai", "Hint", "LaCroix", "Perrier", "San Pellegrino",
            "Canada Dry", "Schweppes", "7UP", "Fanta", "Sunkist", "Barq's Root Beer",
            "A&W Root Beer", "Mug Root Beer", "Vernors", "Squirt", "Fresca",
            "Crystal Light", "Country Time", "Tang", "Kool-Aid", "Hawaiian Punch",
            "Ocean Spray", "Welch's", "V8", "Naked Juice", "Bolthouse Farms",
            "Odwalla", "Suja", "Evolution Fresh", "Coconut Water"
        ],
        "Snacks": [
            "Lay's Potato Chips", "Doritos", "Cheetos", "Pringles", "Fritos",
            "Tostitos", "Ruffles", "Sun Chips", "Popchips", "Gardetto's",
            "Snyder's of Hanover", "Utz", "Kettle Brand", "Boulder Canyon",
            "Herr's", "Wise", "Andy Capp's", "Funyuns", "Bugles", "Pirate's Booty",
            "Smartfood", "Chex Mix", "Munchies", "Combos", "Rold Gold Pretzels",
            "Snyder's Pretzels", "Utz Pretzels", "Mister Bee", "Charles Chips",
            "Zapp's", "Tim's Cascade", "Deep River", "Bare", "Boomchickapop",
            "SkinnyPop", "Angie's Boomchickapop", "Lesser Evil", "Sensible Portions",
            "Popcorners", "Bobby's", "Good Health", "Way Better", "Harvest Snaps",
            "Biena", "Saffron Road", "Beanitos", "Quest", "KIND", "CLIF"
        ],
        "Cookies & Crackers": [
            "Oreo Cookies", "Chips Ahoy", "Nilla Wafers", "Teddy Grahams",
            "Ritz Crackers", "Wheat Thins", "Triscuit", "Cheez-It", "Goldfish",
            "Saltines", "Graham Crackers", "Barnum's Animals", "Pepperidge Farm",
            "Keebler", "Nabisco", "Belvita", "Carr's", "Stoned Wheat Thins",
            "Social Tea", "Animal Crackers", "Girl Scout Cookies", "Mrs. Fields",
            "Chips Deluxe", "Famous Amos", "Archway", "Little Debbie", "Hostess",
            "Drake's", "Entenmann's", "Tastykake", "LU", "Biscoff", "Walkers Shortbread",
            "McVitie's", "Milano", "Lorna Doone", "Fig Newton", "Oatmeal Creme Pie",
            "Fudge Stripes", "Coconut Dreams", "Vanilla Wafers", "Biscotti",
            "Stella D'Oro", "Waverly", "Sociables", "Club Crackers", "Town House",
            "Bremner", "Premium", "Kashi", "Annie's"
        ],
        "Candy & Chocolate": [
            "M&M's", "Snickers", "Twix", "Milky Way", "3 Musketeers",
            "Reese's Peanut Butter Cups", "Kit Kat", "Hershey's Chocolate Bar",
            "Skittles", "Starburst", "Jolly Rancher", "Jelly Belly", "Haribo",
            "Albanese", "See's Candies", "Godiva", "Lindt", "Ghirardelli",
            "Ferrero Rocher", "Toblerone", "Nestle Crunch", "Butterfinger",
            "Baby Ruth", "100 Grand", "PayDay", "Heath", "Almond Joy", "Mounds",
            "York Peppermint Pattie", "Junior Mints", "Mike and Ike", "Hot Tamales",
            "Sour Patch Kids", "Swedish Fish", "Twizzlers", "Red Vines", "Now and Later",
            "Laffy Taffy", "Airheads", "Nerds", "Runts", "Smarties", "SweetTarts",
            "Tootsie Roll", "Tootsie Pops", "Dots", "Chuckles", "Charleston Chew",
            "Clark Bar", "Zagnut", "Valomilk", "Goetze's", "Brach's", "Russell Stover"
        ],
        "Breakfast Cereals": [
            "Cheerios", "Honey Nut Cheerios", "Frosted Flakes", "Special K",
            "Raisin Bran", "Corn Flakes", "Lucky Charms", "Cap'n Crunch",
            "Fruit Loops", "Apple Jacks", "Cocoa Puffs", "Trix", "Kix",
            "Life", "Grape Nuts", "Shredded Wheat", "Total", "Product 19",
            "Honey Bunches of Oats", "Cinnamon Toast Crunch", "Cookie Crisp",
            "Count Chocula", "Franken Berry", "Boo Berry", "Alpha-Bits",
            "Apple Cinnamon Cheerios", "Multi Grain Cheerios", "Berry Burst Cheerios",
            "Fruity Pebbles", "Cocoa Pebbles", "Dinosaur Eggs", "Golden Grahams",
            "Smacks", "Cracklin' Oat Bran", "Mueslix", "Just Bunches", "Malt-O-Meal",
            "Great Grains", "Fiber One", "All-Bran", "Bran Flakes", "40% Bran Flakes",
            "Complete Wheat Bran", "Oatmeal Crisp", "Harvest Crunch", "Quaker Oats",
            "Quisp", "Quake", "Puffa Puffa Rice", "King Vitaman", "Kaboom"
        ],
        "Breakfast Foods": [
            "Pop Tarts", "Eggo Waffles", "Toaster Strudel", "Jimmy Dean Sausage",
            "Oscar Mayer Bacon", "Hillshire Farm", "Bob Evans", "Jones Dairy Farm",
            "Aunt Jemima", "Mrs. Butterworth's", "Log Cabin", "Hungry Jack",
            "Krusteaz", "Bisquick", "Pillsbury", "Thomas'", "Pepperidge Farm Bread",
            "Sara Lee", "Wonder Bread", "Nature's Own", "Arnold", "Orowheat",
            "Entenmann's", "Hostess", "Little Debbie", "Duncan Hines", "Betty Crocker",
            "Jiffy", "Kashi", "Cascadian Farm", "Udi's", "Canyon Bakehouse",
            "Rudi's", "Dave's Killer Bread", "Ezekiel", "Food for Life",
            "Vans", "Kodiak Cakes", "Birch Benders", "Arrowhead Mills",
            "Bob's Red Mill", "King Arthur", "Gold Medal", "Robin Hood",
            "Hodgson Mill", "Red Star", "Fleischmann's", "Jimmy Dean Breakfast Bowls",
            "Farm Rich", "Van's"
        ],
        "Dairy & Eggs": [
            "Yoplait Yogurt", "Chobani Yogurt", "Dannon Yogurt", "Activia",
            "Oikos", "Fage", "Siggi's", "Noosa", "Land O'Lakes Butter",
            "Challenge Butter", "Tillamook", "Kerrygold", "Organic Valley",
            "Horizon Organic", "Califia Farms", "Silk", "So Delicious",
            "Almond Breeze", "Pacific Foods", "Oatly", "Chobani Oat",
            "Breakstone's Sour Cream", "Daisy", "Kemps", "Dean Foods",
            "Borden", "Prairie Farms", "Garelick Farms", "Hood", "Cabot",
            "Crystal Farms", "Lactaid", "Fairlife", "Ultra-Filtered Milk",
            "Eggland's Best", "Nellie's Free Range", "Pete and Gerry's",
            "Vital Farms", "Happy Egg", "Organic Valley Eggs", "Kirkland",
            "Great Value", "365", "Market Pantry", "Simple Truth",
            "Clover Sonoma", "Straus Family Creamery", "Stonyfield",
            "Brown Cow", "Wallaby", "Nancy's"
        ],
        "Frozen Foods": [
            "Hot Pockets", "Lean Cuisine", "Stouffer's", "Digiorno Pizza",
            "Tombstone Pizza", "Red Baron", "California Pizza Kitchen",
            "Birds Eye Vegetables", "Green Giant", "Marie Callender's",
            "Banquet", "Healthy Choice", "Amy's Kitchen", "Evol", "Smart Ones",
            "Bertolli", "Michelina's", "Celentano", "Gorton's", "Mrs. Paul's",
            "Van de Kamp's", "SeaPak", "Trident", "High Liner", "Fisher Boy",
            "Kid Cuisine", "Tony's", "Jeno's", "Totino's", "Ellio's",
            "Home Run Inn", "Screamin' Sicilian", "Freschetta", "Talia di Napoli",
            "Kashi Pizza", "Caulipower", "Against the Grain", "Udi's Pizza",
            "Sweet Earth", "Daily Harvest", "Tattooed Chef", "Modern Table",
            "Bubba Burgers", "Dr. Praeger's", "MorningStar Farms", "Gardein",
            "Beyond Meat", "Impossible Foods", "Tyson", "Perdue"
        ],
        "Canned & Packaged Goods": [
            "Campbell's Soup", "Progresso Soup", "Chef Boyardee", "SpaghettiOs",
            "Hormel Chili", "Dinty Moore Stew", "Spam", "Bush's Beans",
            "Del Monte", "Libby's", "Hunt's", "Muir Glen", "Green Giant",
            "B&M", "Allens", "S&W", "Le Sueur", "Goya", "La Preferida",
            "Rosarita", "Old El Paso", "Ortega", "Chi-Chi's", "Hormel Compleats",
            "Chef Francisco", "Wolf Brand Chili", "Nalley", "Van Camp's",
            "Margaret Holmes", "Glory Foods", "Trappey's", "Luck's",
            "Hormel Mary Kitchen", "Hormel Corned Beef", "Underwood",
            "Bumble Bee", "Chicken of the Sea", "Starkist", "Wild Planet",
            "Safe Catch", "Crown Prince", "King Oscar", "Season", "Annie's",
            "Amy's", "Pacific Foods", "Imagine", "Health Valley", "Eden Foods"
        ],
        "Condiments & Sauces": [
            "Heinz Ketchup", "French's Mustard", "Hellmann's Mayonnaise",
            "Hidden Valley Ranch", "Wish-Bone Salad Dressing", "Kraft",
            "Sweet Baby Ray's BBQ", "Stubb's", "Annie's", "Sir Kensington's",
            "Primal Kitchen", "Grey Poupon", "Gulden's", "Plochman's",
            "Boar's Head", "Best Foods", "Duke's", "Blue Plate", "Kewpie",
            "Miracle Whip", "Ken's", "Newman's Own", "Brianna's", "Marzetti",
            "Girard's", "Cardini's", "Lawry's", "A.1. Steak Sauce",
            "Lea & Perrins", "Texas Pete", "Tabasco", "Cholula", "Sriracha",
            "Frank's RedHot", "Crystal", "Louisiana", "Tapatio", "Valentina",
            "Pickapeppa", "Pickle Juice", "Claussen", "Vlasic", "Mt. Olive",
            "Wickles", "Grillo's", "Bubbies", "McClure's", "Ba-Tampte"
        ],
        "Baking & Cooking": [
            "Pam Cooking Spray", "Crisco", "Wesson Oil", "Bertolli",
            "Filippo Berio", "King Arthur Flour", "Gold Medal", "Domino Sugar",
            "Imperial Sugar", "Tate & Lyle", "McCormick", "Spice Islands",
            "Diamond Crystal", "Morton", "C&H", "Florida Crystals", "Wholesome",
            "Swerve", "Lakanto", "Truvia", "Stevia", "Splenda", "Equal",
            "Sweet'N Low", "Pyure", "Swanson Broth", "College Inn",
            "Kitchen Basics", "Better Than Bouillon", "Herb-Ox", "Wyler's",
            "Knorr", "Maggi", "Orrington Farms", "Frontier Co-op", "Simply Organic",
            "Penzey's", "The Spice House", "Badia", "Goya Adobo", "Sazon",
            "Tony Chachere's", "Slap Ya Mama", "Cavender's", "Everglades Seasoning",
            "Lawry's Seasoned Salt", "Johnny's", "Kikkoman", "San-J", "Lee Kum Kee"
        ],
        "Coffee & Tea": [
            "Folgers Coffee", "Maxwell House", "Starbucks Coffee", "Nescafe",
            "Keurig", "Green Mountain", "Peet's Coffee", "Lavazza",
            "Illy", "Tazo", "Twinings", "Bigelow", "Celestial Seasonings",
            "Yogi Tea", "Traditional Medicinals", "Community Coffee",
            "Chock full o'Nuts", "Eight O'Clock", "Cafe Bustelo", "Cafe Pilon",
            "Cafe Llave", "Medaglia d'Oro", "Cafe du Monde", "Death Wish Coffee",
            "Black Rifle Coffee", "Kicking Horse", "Intelligentsia", "Stumptown",
            "Blue Bottle", "Counter Culture", "La Colombe", "Philz", "Verena Street",
            "New England Coffee", "Gloria Jean's", "The Coffee Bean & Tea Leaf",
            "David's Tea", "Numi", "Mighty Leaf", "Republic of Tea",
            "Harney & Sons", "Mariage Freres", "Dilmah", "Tetley", "PG Tips",
            "Barry's Tea", "Lyon's", "Bewley's", "Nambarrie", "Typhoo"
        ],
        "Health & Nutrition": [
            "Clif Bar", "Luna Bar", "Kind Bar", "Larabar", "RXBAR",
            "Quest", "Atkins", "ZonePerfect", "ThinkThin", "PowerBar",
            "Gatorade Protein", "Premier Protein", "Ensure", "Boost",
            "Glucerna", "Carnation Breakfast Essentials", "SlimFast",
            "Optifast", "Medifast", "Isagenix", "Shakeology", "Vega",
            "Garden of Life", "Orgain", "OWYN", "Soylent", "Huel",
            "Ample", "Kachava", "Lyfe Fuel", "Transparent Labs",
            "Ghost", "MuscleTech", "Optimum Nutrition", "BSN", "Cellucor",
            "Dymatize", "Isopure", "GNC", "Nature's Bounty", "Nature Made",
            "Solgar", "NOW Foods", "Jarrow Formulas", "Country Life",
            "Garden of Life vitamins", "Ritual", "Care/of", "Persona", "HUM"
        ],
        "Natural & Organic": [
            "Annie's Homegrown", "Nature's Path", "Kashi", "Amy's Kitchen",
            "Stonyfield", "Earthbound Farm", "Cascadian Farm", "Muir Glen",
            "Newman's Own", "Bob's Red Mill", "Arrowhead Mills", "Eden Foods",
            "Whole Foods 365", "Simple Truth", "O Organics", "Open Nature",
            "Signature SELECT", "Market Pantry", "Good & Gather", "Great Value Organic",
            "Kirkland Organic", "Happy Baby", "Plum Organics", "Earth's Best",
            "Gerber Organic", "Beech-Nut Organic", "Sprout", "Once Upon a Farm",
            "Yummy Spoonfuls", "NurturMe", "Little Spoon", "Raised Real",
            "Thrive Market", "Misfits Market", "Imperfect Foods", "Hungryroot",
            "Daily Harvest", "Sun Basket", "HelloFresh", "Blue Apron",
            "Green Chef", "Home Chef", "EveryPlate", "Dinnerly", "Marley Spoon",
            "Purple Carrot", "Sakara", "Veestro", "Factor", "Freshly"
        ],
        "Cosmetics & Personal Care": [
            "Dove", "Olay", "Neutrogena", "Cetaphil", "CeraVe", "Aveeno",
            "Eucerin", "Aquaphor", "Vaseline", "Nivea", "Lubriderm",
            "Jergens", "Gold Bond", "St. Ives", "Burt's Bees", "The Body Shop",
            "L'Occitane", "Kiehl's", "Clinique", "Estee Lauder", "La Mer",
            "SK-II", "Shiseido", "Sulwhasoo", "Drunk Elephant", "Sunday Riley",
            "The Ordinary", "Paula's Choice", "Dr. Barbara Sturm", "Augustinus Bader",
            "Biologique Recherche", "Valmont", "Sisley", "Guerlain", "Chanel",
            "Dior", "Yves Saint Laurent", "Giorgio Armani", "Tom Ford",
            "MAC", "Bobbi Brown", "NARS", "Urban Decay", "Too Faced",
            "Tarte", "Fenty Beauty", "Glossier", "Milk Makeup", "Kosas",
            "Ilia", "RMS Beauty", "Westman Atelier", "Jones Road", "Merit"
        ],
        "Cleaning & Household": [
            "Clorox", "Lysol", "Mr. Clean", "Dawn", "Palmolive", "Ajax",
            "Comet", "Soft Scrub", "Bar Keepers Friend", "Method", "Seventh Generation",
            "Mrs. Meyer's Clean Day", "Better Life", "Bon Ami", "Biokleen",
            "Ecover", "Attitude", "Branch Basics", "Force of Nature",
            "Grove Collaborative", "Blueland", "Dropps", "Tru Earth",
            "Earth Breeze", "Nellie's", "Charlie's Soap", "Dr. Bronner's",
            "Molly's Suds", "J.R. Watkins", "Meyer's", "Puracy", "Common Good",
            "Babyganics", "The Honest Company", "Dapple", "Mustela", "Cetaphil Baby",
            "Aveeno Baby", "Johnson's Baby", "Burts Bees Baby", "Earth Mama",
            "California Baby", "Weleda", "Babo Botanicals", "Hello Bello"
        ],
        "Pet Food": [
            "Purina", "Pedigree", "Iams", "Hill's Science Diet", "Royal Canin",
            "Blue Buffalo", "Merrick", "Wellness", "Natural Balance", "Taste of the Wild",
            "Fromm", "Orijen", "Acana", "Instinct", "Stella & Chewy's",
            "Primal", "Steve's Real Food", "Smallbatch", "Answers", "The Honest Kitchen",
            "Farmina", "Nulo", "Ziwi Peak", "K9 Natural", "Feline Natural",
            "Weruva", "Tiki Cat", "Dave's Pet Food", "Solid Gold", "Nutro",
            "Rachael Ray Nutrish", "Authority", "Simply Nourish", "Wholehearted",
            "American Journey", "Wild Earth", "Jiminy's", "Bond Pet Foods",
            "Because Animals", "Wild One", "Spot & Tango", "The Farmer's Dog",
            "Ollie", "Nom Nom", "Just Food For Dogs", "A Pup Above"
        ],
        "Baby Food": [
            "Gerber", "Beech-Nut", "Earth's Best", "Happy Baby", "Plum Organics",
            "Once Upon a Farm", "Yummy Spoonfuls", "NurturMe", "Little Spoon",
            "Raised Real", "Sprout", "Cerebelly", "Amara", "Nurture Life",
            "Tiny Organics", "Serenity Kids", "Mushies", "Bambinos",
            "Baby Gourmet", "Ella's Kitchen", "Hipp", "Holle", "Lebenswert",
            "Topfer", "Bioland", "Demeter", "Nestle Baby", "Parent's Choice",
            "Up & Up", "Good & Gather", "Simple Truth", "Kirkland", "365"
        ]
    }

    # Create DataFrame using the template structure
    data = []

    for category, brands in categories_data.items():
        for brand in brands:
            data.append({
                "Product_Brand": brand,
                "Category": category,
                "Notes": "",
                "B_Corp": False,
                "Fair_Trade": False,
                "Rainforest_Alliance": False,
                "Leaping_Bunny": False,
                "Research_Complete": False,
                "Last_Updated": "",
                "Confidence": "Low"
            })

    df = pd.DataFrame(data)

    # Sort by Category then Product_Brand
    df = df.sort_values(["Category", "Product_Brand"]).reset_index(drop=True)

    # Create Excel file
    output_file = "comprehensive_grocery_certifications.xlsx"

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Certifications', index=False)

        # Get workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Certifications']

        # Apply formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Format headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Adjust column widths
        column_widths = {
            'A': 35,  # Product_Brand
            'B': 25,  # Category
            'C': 30,  # Notes
            'D': 10,  # B_Corp
            'E': 12,  # Fair_Trade
            'F': 20,  # Rainforest_Alliance
            'G': 15,  # Leaping_Bunny
            'H': 18,  # Research_Complete
            'I': 15,  # Last_Updated
            'J': 12   # Confidence
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Apply borders to all cells
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, max_col=10):
            for cell in row:
                cell.border = thin_border
                if cell.column in [4, 5, 6, 7, 8]:  # Boolean and date columns
                    cell.alignment = Alignment(horizontal="center")

        # Freeze header row
        worksheet.freeze_panes = 'A2'

    print(f"Excel file created: {output_file}")
    print(f"Total products: {len(df)}")
    print(f"Categories: {len(categories_data)}")

    # Print summary by category
    print("\nProducts per category:")
    category_counts = df['Category'].value_counts()
    for category, count in category_counts.items():
        print(f"{category}: {count} products")

    return df

def add_statistics_sheet():
    """
    Add a statistics sheet to the Excel file with category summaries
    """
    df = create_comprehensive_grocery_sheet()
    output_file = "comprehensive_grocery_certifications.xlsx"

    # Load the workbook to add a statistics sheet
    workbook = openpyxl.load_workbook(output_file)

    # Create statistics sheet
    if 'Statistics' in workbook.sheetnames:
        del workbook['Statistics']

    stats_sheet = workbook.create_sheet('Statistics')

    # Calculate statistics
    category_stats = df['Category'].value_counts().reset_index()
    category_stats.columns = ['Category', 'Product Count']

    # Add certification column counts
    certification_columns = ['B_Corp', 'Fair_Trade', 'Rainforest_Alliance', 'Leaping_Bunny']
    for cert in certification_columns:
        cert_stats = df.groupby('Category')[cert].apply(lambda x: (x == True).sum()).reset_index()
        cert_stats.columns = ['Category', f'{cert}_True']
        category_stats = category_stats.merge(cert_stats, on='Category', how='left')

    # Write statistics to sheet
    headers = list(category_stats.columns)
    stats_sheet.append(headers)

    for _, row in category_stats.iterrows():
        stats_sheet.append(list(row))

    # Format statistics sheet
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F6228", end_color="4F6228", fill_type="solid")

    for cell in stats_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        stats_sheet.column_dimensions[col_letter].width = 20

    # Add summary
    stats_sheet.append([])
    stats_sheet.append(["Summary Statistics"])
    stats_sheet.append([f"Total Products: {len(df)}"])
    stats_sheet.append([f"Total Categories: {len(category_stats)}"])
    stats_sheet.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])

    workbook.save(output_file)
    print(f"\nStatistics sheet added to {output_file}")

    return df

if __name__ == "__main__":
    print("Creating comprehensive grocery certifications database...")
    print("=" * 60)

    df = add_statistics_sheet()

    print("\n" + "=" * 60)
    print("File creation complete!")
    print("\nNext steps:")
    print("1. Open the Excel file: comprehensive_grocery_certifications.xlsx")
    print("2. Research each brand's certifications")
    print("3. Update the True/False values in columns D-G as you verify certifications")
    print("4. Mark Research_Complete as True when done")
    print("5. Update Last_Updated date")
    print("6. Adjust Confidence level based on source reliability")
    print("\nTemplate maintained: All certifications default to False for manual verification")
